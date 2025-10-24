"""
XLSX to CSV Converter
Converts Microsoft Excel spreadsheets to CSV format preserving sheets and data
"""
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any
from openpyxl import load_workbook
import csv
import io


def convert(file_path: Path, output_dir: Path) -> Dict[str, Any]:
    """
    Convert XLSX file to CSV format (one CSV per sheet)
    
    Args:
        file_path: Path to the input XLSX file
        output_dir: Directory where output should be written
    
    Returns:
        Dictionary with conversion result
    """
    try:
        # Create xlsx output directory if it doesn't exist
        xlsx_dir = output_dir / "xlsx"
        xlsx_dir.mkdir(exist_ok=True)

        # Load the workbook
        wb = load_workbook(str(file_path), read_only=True)

        conversion_metadata = {
            'sheets': [],
            'total_rows': 0,
            'total_sheets': len(wb.sheetnames)
        }

        # Track used output filenames to avoid collisions
        used_output_filenames = set()

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            

            # Skip empty sheets using calculate_dimension()
            dim = ws.calculate_dimension()
            if dim == "A1":
                if ws["A1"].value is None:
                    continue
            
            # Read sheet data
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(row)
            
            if not rows:
                continue
            
            # Determine headers (use first non-empty row)
            headers = None
            header_row_idx = 0
            for i, row in enumerate(rows):
                if any(cell is not None and str(cell).strip() for cell in row):
                    headers = [str(cell) if cell is not None else f"Column_{j+1}" for j, cell in enumerate(row)]
                    header_row_idx = i
                    break
            
            if headers is None:
                continue
            
            # Remove header row from data if it was found
            data_rows = rows[header_row_idx + 1:] if header_row_idx < len(rows) else []
            
            # Limit to 50,000 rows per sheet
            data_rows = data_rows[:50000]
            
            # Create CSV content
            output_buffer = io.StringIO()
            writer = csv.writer(output_buffer, quoting=csv.QUOTE_ALL)
            
            # Write headers
            writer.writerow(headers)
            
            # Write data rows (convert None to empty string, handle different data types)
            for row in data_rows:
                csv_row = []
                for cell in row:
                    if cell is None:
                        csv_row.append("")
                    elif isinstance(cell, (int, float, bool)):
                        csv_row.append(str(cell))
                    else:
                        csv_row.append(str(cell))
                writer.writerow(csv_row)
            
            # Generate output filename with sheet name
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            if not safe_sheet_name:
                safe_sheet_name = f"Sheet_{len(conversion_metadata['sheets']) + 1}"
            

            # Ensure output_filename is unique
            base_output_filename = f"{file_path.stem}_{safe_sheet_name}.csv"
            output_filename = base_output_filename
            suffix = 1
            while output_filename in used_output_filenames:
                output_filename = f"{file_path.stem}_{safe_sheet_name}_{suffix}.csv"
                suffix += 1
            used_output_filenames.add(output_filename)
            output_file = xlsx_dir / output_filename
            
            # Write CSV file
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                f.write(output_buffer.getvalue())
            
            # Add sheet metadata
            sheet_metadata = {
                'name': sheet_name,
                'rows': len(data_rows) + 1,  # +1 for header
                'columns': len(headers),
                'hidden': ws.sheet_state != 'visible'  # Check if sheet is hidden
            }
            conversion_metadata['sheets'].append(sheet_metadata)
            conversion_metadata['total_rows'] += len(data_rows) + 1
        
        # Close workbook
        wb.close()
        
        return {
            'status': 'success',
            'output': str(xlsx_dir.relative_to(output_dir.parent)),
            'converter': 'xlsx_to_csv',
            'metadata': conversion_metadata
        }
        
    except Exception as e:
        return {
            'status': 'failed',
            'error': str(e),
            'converter': 'xlsx_to_csv'
        }
